from typing import List, Dict
import logging

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.geocoding import geocode_address
from app.db.session import SessionLocal

from app.core.config import settings
from app.crud.crud_location import bulk_insert_locations


logger = logging.getLogger(settings.PROJECT_NAME)


async def serialize_excel(
    spreadsheet: Worksheet
) -> Dict:
    locations = []
    unprocessed_locations = []
    for row in spreadsheet.iter_rows(values_only=True):
        if not row[0]:
            continue
        try:
            location = {
                "address": row[0].strip(),
                "street_number": row[1],
                "city": row[2],
                "country": row[3],
                "postcode": row[4],
                "reports": {
                    "buildingCondition": {
                        "flag": row[5].lower(),
                        "description": row[11] if row[11] else ""
                    },
                    "electricity": {
                        "flag": row[6].lower(),
                        "description": ""
                    },
                    "carEntrance": {
                        "flag": row[7].lower(),
                        "description": ""
                    },
                    "water": {
                        "flag": row[8].lower(),
                        "description": ""
                    },
                    "fuelStation": {
                        "flag": row[9].lower(),
                        "description": ""
                    },
                    "hospital": {
                        "flag": row[10].lower(),
                        "description": ""
                    }
                }
            }
            locations.append(location)

        except Exception as e:
            unprocessed_locations.append({
                "location": row,
                "code": "SERIALIZATION_ERROR",
                "detail": e
            })

    return {
        "processed": locations,
        "unprocessed": unprocessed_locations
    }


async def geocode_locations(
    locations: List[Dict]
) -> Dict:

    geocoded_locations = []
    unprocessed_locations = []
    for location in locations:
        address_string = "{}, {}".format(location.get('address'), location.get('street_number'))
        coordinates = geocode_address(
            address=address_string,
            city=location.get('city')
        )
        if not coordinates:
            unprocessed_locations.append({
                "location": location,
                "code": "GEOCODING_ERROR",
                "detail": "Address not found"
            })
            continue
        location["lat"] = coordinates.latitude
        location["lng"] = coordinates.longitude
        geocoded_locations.append(location)

    return {
        "geocoded": geocoded_locations,
        "unprocessed": unprocessed_locations
    }


async def upload_locations(
        filepath: str,
        doctype: str
):

    unprocessed_locations = []

    if doctype == "excel":
        workbook = load_workbook(filepath)
        sheet = workbook.active
        logger.debug("XLSX loaded, starting serialization.")

        serialization_results = await serialize_excel(sheet)
        unprocessed_locations = serialization_results.get('unprocessed')
        serialized_locations = serialization_results.get('processed')
        logger.debug("Processed locations: {}".format(len(serialized_locations)))
        logger.debug("Unprocessed locations: {}".format(len(unprocessed_locations)))

        logger.debug("Starting geocoding")
        geocoding_results = await geocode_locations(serialized_locations)
        geocoded_locations = geocoding_results.get('geocoded')
        unprocessed_locations.append(loc for loc in geocoding_results.get('unprocessed'))
        logger.debug("Geocoded locations: {}".format(len(geocoded_locations)))
        logger.debug("Could not geocode: {}".format(len(geocoding_results.get('unprocessed'))))

        logger.debug("Adding {} locations to database".format(len(geocoded_locations)))
        db_locations = bulk_insert_locations(
            db=SessionLocal(),
            locations=geocoded_locations
        )
        added_locations = db_locations.get("added")
        unprocessed_locations.append(loc for loc in db_locations.get('unprocessed'))
        logger.debug("Added locations: {}".format(len(added_locations)))
        logger.debug("Failed to add: {}".format(len(db_locations.get('unprocessed'))))

        return unprocessed_locations

